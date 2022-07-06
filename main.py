import csv
import io
from bs4 import BeautifulSoup
import requests

def main():
    import openpyxl

    wb = openpyxl.load_workbook('komus_opt_ret_4.xlsx')
    # wb = openpyxl.load_workbook('test.xlsx')
    ws = wb['Опт прайс-лист']
    cnt = 0
    print("Loaded!")
    for col in range(12071, 12172):
        try:

            url = ws.cell(row=col, column=5).hyperlink.target
            # print(url)
            request = requests.get(url)
            page = BeautifulSoup(request.text, features='lxml')
            description = page.find("div", "tab-pane").text
            ws.cell(row=col, column=24).value = description

        except Exception:
            pass
    wb.save('data6.xlsx')
    wb.close()

if __name__ == '__main__':
    main()